# Implementation of Selenium WebDriver with Python using PyTest
import unittest
import pytest
from selenium import webdriver
import sys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.by import By
from time import sleep
import HtmlTestRunner
# importamos el submodulo "Workbook"
from openpyxl import load_workbook


def test_condominio():
    chrome_driver = webdriver.Chrome()

    chrome_driver.get('https://prueba.ant.cl/mi-portal/Login')
    chrome_driver.maximize_window()

    # especificamos el nombre y la ruta del archivo
    filesheet = "..\Datos\Datos_Condominio.xlsx"

    # creamos el obejeto load_workbook
    wb = load_workbook(filesheet)

    # seleccionamos la Hoja del archivo
    sheet = wb['User']

    # Obtenemos el valor de la celda A1
    email = sheet['A2'].value
    passw = sheet['B2'].value

    wait = WebDriverWait(chrome_driver, 2)
    wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//input[@id='login-password']")))

    elem = chrome_driver.find_element_by_xpath("//input[@id='login-user-name']")
    elem.clear()
    elem.send_keys(email)
    # elem.send_keys(Keys.RETURN)
    elem = chrome_driver.find_element_by_xpath("//input[@id='login-password']")
    elem.clear()
    #    elem.send_keys("Fz8yV)TT&hfr")
    elem.send_keys(passw)
    #    sleep(5)
    elem.send_keys(Keys.RETURN)

    # Voy a pinchar opción Ramo Wizard en el menú
    # open tab
    #chrome_driver.find_element_by_tag_name('body').send_keys(Keys.CONTROL + 't')
    #chrome_driver.get('https://prueba.ant.cl/mi-portal/app/posts')
    #elem = chrome_driver.find_element_by_xpath('//span[@class="hover:wi-border-b-3 wi-border-purple-400 wi-px-2"]')


    # Voy a pinchar opción Mantenedores en el menú
    wait = WebDriverWait(chrome_driver, 15)
    wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//span[text()='Mantenedores']")))
    elem = chrome_driver.find_element_by_xpath("//span[text()='Mantenedores']").click()

    # Acá llama a la opción Ramo Wizard pero con menú extendido
    elem = chrome_driver.find_element_by_xpath("//a[text()='Condominio']").click()

    #sleep(10)
    timeout = 15
    try:
        element_present = expected_conditions.presence_of_element_located((By.ID, 'iframe-render'))
        WebDriverWait(chrome_driver, timeout).until(element_present)
    except TimeoutException:
        print('Timed out waiting for page to load')

    # Me cambio al iframe de ingreso de datos
    iframe = chrome_driver.find_element_by_id("iframe-render")
    chrome_driver.switch_to.frame(iframe)

    #wait = WebDriverWait(chrome_driver, 25)
    #wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//input[@id='PerAsegurado_Identificacion']")))

    # seleccionamos la Hoja del archivo con los datos
    sheet = wb['Datos CP 1']

    RutCondominio = sheet['B2'].value
    RazonSocial = sheet['B2'].value
    NroReserva = sheet['C2'].value
    CHUBB = sheet['D2'].value
    Renta = sheet['E2'].value
    Comuna = sheet['F2'].value
    Direccion = sheet['G2'].value
    ConstruccionMuro = sheet['H2'].value
    ConstTecho = sheet['I2'].value
    TipoCondominio = sheet['J2'].value
    Antiguedad = sheet['K2'].value
    UbicacionRiesgo = sheet['L2'].value

    elem = chrome_driver.find_element_by_id('PerAsegurado_Identificacion')
    elem.clear()
    elem.send_keys(RutCondominio)
    elem.send_keys(Keys.RETURN)

    elem = chrome_driver.find_element_by_xpath("//input[@id='NReservaBCI_Texto']")
    elem.clear()
    elem.send_keys(NroReserva)
    elem.send_keys(Keys.RETURN)

    sleep(5)

    elem = chrome_driver.find_element_by_id('PerAsegurado_RazonSocial')
    #('PerAsegurado_RazonSocial')
    #    elem.send_keys(Keys.RETURN)
    assert elem.get_attribute('value') == RazonSocial

    sleep(5)


    chrome_driver.close()
		

if __name__ == "__main__":
#if __name__ == '__main__':
    unittest.main(testRunner=HtmlTestRunner.HTMLTestRunner(output='D:/MisCosas/Desarrollo/Testing_Seguros/reports'))
	