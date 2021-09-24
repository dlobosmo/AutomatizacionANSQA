# Implementation of Selenium WebDriver with Python using PyTest
import unittest
import pytest
from selenium import webdriver
import sys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.by import By
from time import sleep
import HtmlTestRunner
# importamos el submodulo "Workbook"
from openpyxl import load_workbook


def test_condominio():
    chrome_driver = webdriver.Chrome()

    chrome_driver.get('https://prueba.ant.cl/mi-portal/Login')
    # No quitar esta opción, ya que minimizado cambia todos los componentes siguientes.
    chrome_driver.maximize_window()

    # especificamos el nombre y la ruta del archivo
    filesheet = "..\Datos\Datos_Condominio.xlsx"

    # creamos el objeto load_workbook
    wb = load_workbook(filesheet)

    # seleccionamos la Hoja del archivo
    sheet = wb['User']

    # Obtenemos el valor de la celda A2 y B2
    email = sheet['A2'].value
    passw = sheet['B2'].value

    wait = WebDriverWait(chrome_driver, 2)
    wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//input[@id='login-password']")))

    elem = chrome_driver.find_element_by_xpath("//input[@id='login-user-name']")
    elem.clear()
    elem.send_keys(email)
    elem = chrome_driver.find_element_by_xpath("//input[@id='login-password']")
    elem.clear()
    elem.send_keys(passw)
    elem.send_keys(Keys.RETURN)


    # Voy a pinchar opción Mantenedores en el menú
    wait = WebDriverWait(chrome_driver, 15)
    wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//span[text()='Mantenedores']")))
    elem = chrome_driver.find_element_by_xpath("//span[text()='Mantenedores']").click()

    # Acá llama a la opción Ramo Wizard pero con menú extendido
    elem = chrome_driver.find_element_by_xpath("//a[text()='Ramo Wizard']").click()

    timeout = 15
    try:
        element_present = expected_conditions.presence_of_element_located((By.ID, 'iframe-render'))
        WebDriverWait(chrome_driver, timeout).until(element_present)
    except TimeoutException:
        print('Tiempo excedido para cargar la página')

    # Me cambio al iframe de ingreso de datos
    iframe = chrome_driver.find_element_by_id("iframe-render")
    chrome_driver.switch_to.frame(iframe)

    # seleccionamos la Hoja del archivo con los datos
    sheet = wb['Datos CP14']

    FormadePago = sheet['B2'].value
    Filtrobuscar = sheet['C2'].value

    #Voy a editar RAMO
    wait = WebDriverWait(chrome_driver, 15)
    wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//input[@id='customRadioInline1']")))
    #elem = chrome_driver.find_element_by_id('customRadioInline1').click()
    elem = chrome_driver.find_element_by_xpath("//input[@id='customRadioInline1' and @value='editar']").click()
    #"#customRadioInline1"
    #elem.clear()
    #elem.send_keys(RutCondominio)
    #elem.send_keys(Keys.RETURN)

    #Debemos seleccionar Condominio de una tabla paginada, primero seleccionamos buscar y buscamos Condominio
    #sleep(5)
    wait = WebDriverWait(chrome_driver, 10)
    wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//input[@aria-controls='tblListaRamos']")))
    #elem = chrome_driver.find_element_by_xpath("//input[@type='search' and @aria-controls='tblListaRamos']").click()
    elem = chrome_driver.find_element_by_xpath("//input[@aria-controls='tblListaRamos']")
    elem.clear()
    elem.send_keys(Filtrobuscar)


    #seleccione editar condominio
    elem = chrome_driver.find_element_by_xpath("//*[@id='tblListaRamos']/tbody/tr[1]/td[2]/a/i").click()

    #Selecciono cambio a mostrar o no mostrar forma de pago
    #sleep(10)
    if FormadePago == "S":
        wait = WebDriverWait(chrome_driver, 10)
        wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//*[@id='frmIngresarRamo']/div[2]/div/div/div[2]/div/div[2]/input[1]")))
        #elem = chrome_driver.find_element_by_xpath("//input[@type='radio' and @value='true' and name='MostrarSeccionFormaDePago']").click()
        elem = chrome_driver.find_element_by_xpath("//*[@id='frmIngresarRamo']/div[2]/div/div/div[2]/div/div[2]/input[2]").click()

    if FormadePago == "N":
        wait = WebDriverWait(chrome_driver, 10)
        wait.until(expected_conditions.element_to_be_clickable(
            (By.XPATH, "//*[@id='frmIngresarRamo']/div[2]/div/div/div[2]/div/div[2]/input[1]")))
        # elem = chrome_driver.find_element_by_xpath("//input[@type='radio' and @value='false' and name='MostrarSeccionFormaDePago']").click()
        elem = chrome_driver.find_element_by_xpath(
            "//*[@id='frmIngresarRamo']/div[2]/div/div/div[2]/div/div[2]/input[1]").click()

    # Voy a botón Continuar para grabar el cambio
    elem = chrome_driver.find_element_by_xpath("//*[@id='ramoParte0']/div[1]/div[5]/div/a").click()

    #Valido que el cambio fue realizado...
    wait = WebDriverWait(chrome_driver, 30)
    wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//*[@id='ramoParte1']/div/div/div/a[1]")))

    # Voy a pinchar opción Mantenedores en el menú
    # Me cambio al iframe de menú primero
    chrome_driver.switch_to.default_content()
    #Selecciono opción Mantenedores en el menú
    elem = chrome_driver.find_element_by_xpath("//span[text()='Mantenedores']").click()
    # Acá llama a la opción Condominio
    elem = chrome_driver.find_element_by_xpath("//a[text()='Condominio']").click()
    #espera a que cargue  la página con un tiempo máximo de 15 segundos
    timeout = 15
    try:
        element_present = expected_conditions.presence_of_element_located((By.ID, 'iframe-render'))
        WebDriverWait(chrome_driver, timeout).until(element_present)
    except TimeoutException:
        print('Tiempo excedido para cargar la página')

    # Me cambio al iframe de ingreso de datos
    iframe = chrome_driver.find_element_by_id("iframe-render")
    chrome_driver.switch_to.frame(iframe)

    #Se ahce scroll bar para llegar al final de la página y validar si está activo Forma de pago
    chrome_driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    chrome_driver.save_screenshot('..\Screenshot\CP14\Parte03_FormaPagoAparece.png')

    ##Valido que el componente de Forma de pago esté o no presente.
    ##Si FormadePago  = "S" quiere decir que el componente debe existir, en caso contrario debe estar apagado
    #wait = WebDriverWait(chrome_driver, 60)
    #wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//*[@id='frmMateriaAsegurada']/article[4]/div[1]/div/h2']")))
    #if FormadePago == "S":
    #    elem = chrome_driver.find_element_by_xpath("//*[@id='frmMateriaAsegurada']/article[4]/div[1]/div/h2']")
    #    #elem = chrome_driver.find_element_by_id('select2-TipoMedioPago-container')
    #    if elem:
    #        print("OK, se encontró el elemento")
    #        chrome_driver.save_screenshot('..\Screenshot\CP14\Parte03_FormaPagoAparece.png')
    #    else:
    #        print("Error, el elemento no sigue activo")
    #if FormadePago == "N":
    #    try:
    #        elem = chrome_driver.find_element_by_xpath("//*[@id='select2-TipoMedioPago-container']']")
    #        print("Error, el elemento sigue activo")
    #    except AssertionError:
    #        print("OK, No se encontró el elemento")

    ##Cierro sesión
    ##Me cambio al iframe de menú primero
    ##chrome_driver.switch_to.default_content()
    #elem = chrome_driver.find_element_by_xpath("//span[class()='p-avatar-text']").click()
    #elem = chrome_driver.find_element_by_xpath("//span[text()='Cerrar sesión']").click()

    #Cierro navegador
    chrome_driver.close()


if __name__ == "__main__":
    # if __name__ == '__main__':
    unittest.main(testRunner=HtmlTestRunner.HTMLTestRunner(output='D:/MisCosas/Desarrollo/Testing_Seguros/reports'))
