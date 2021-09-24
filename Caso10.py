# Implementation of Selenium WebDriver with Python using PyTest
import unittest
import pytest
from selenium import webdriver
import sys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.by import By
from time import sleep
import HtmlTestRunner
# importamos el submodulo "Workbook"
from openpyxl import load_workbook


def test10_condominio():
    # Especificamos el nombre y la ruta del archivo de datos a leer
    filesheet = "..\Datos\Datos_Condominio.xlsx"

    # Creamos el obejeto load_workbook para leer los datos de excel
    wb = load_workbook(filesheet)

    # seleccionamos la Hoja del archivo con datos de la ruta acceso e inicio de sesión
    sheet = wb['User']
    # obtenemos url de la pagina
    urlPagina = sheet['D2'].value
    # Definicion del chromedriver
    rutaChromeDriver = "..\drivers\chromedriver.exe"
    chrome_driver = webdriver.Chrome(rutaChromeDriver)
    # Driver llama la url
    chrome_driver.get(urlPagina)
    # Se invoca control para maximizar windows
    chrome_driver.maximize_window()


    # Obtenemos el valor de la celda para leer el usuario y contraseña de quien iniciará sesion en el sistema
    email = sheet['A2'].value
    passw = sheet['B2'].value

    wait = WebDriverWait(chrome_driver, 70)
    wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//input[@id='login-password']")))

    # Se ingresa el usuario del sistema
    elem = chrome_driver.find_element_by_xpath("//input[@id='login-user-name']")
    elem.clear()
    elem.send_keys(email)

    # Se ingresa la password del usuario del sistema
    elem = chrome_driver.find_element_by_xpath("//input[@id='login-password']")
    elem.clear()
    elem.send_keys(passw)
    chrome_driver.save_screenshot('..\Screenshot\CP10\Inicio_Sesion.png')
    elem.send_keys(Keys.RETURN)

    # Voy a pinchar opción Mantenedores en el menú
    wait = WebDriverWait(chrome_driver, 60)
    wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//span[text()='Mantenedores']")))
    elem = chrome_driver.find_element_by_xpath("//span[text()='Mantenedores']").click()
    # Acá llama a la opción Ramo Wizard pero con menú extendido
    elem = chrome_driver.find_element_by_xpath("//a[text()='Condominio']").click()
    chrome_driver.save_screenshot('..\Screenshot\CP10\Home_Sistema.png')

    timeout = 30
    try:
        element_present = expected_conditions.presence_of_element_located((By.ID, 'iframe-render'))
        WebDriverWait(chrome_driver, timeout).until(element_present)
    except TimeoutException:
        print('Timed out waiting for page to load')

    # Me cambio al iframe de ingreso de datos
    iframe = chrome_driver.find_element_by_id("iframe-render")
    chrome_driver.switch_to.frame(iframe)

    # seleccionamos la Hoja del archivo con los datos de prueba del caso
    sheet = wb['Datos CP 1']

    RutCondominio = sheet['B11'].value
    RazonSocial = sheet['C11'].value
    NroReserva = sheet['D11'].value
    CONSORCIO = sheet['E11'].value
    CHUBB = sheet['F11'].value
    Renta = sheet['G11'].value
    Direccion = sheet['I11'].value
    ConstruccionMuro = sheet['J11'].value
    ConstTecho = sheet['K11'].value
    TipoCondominio = sheet['L11'].value
    NPisos = sheet['M11'].value
    NSubterraneos = sheet['N11'].value
    BienesEspacio = sheet['Q11'].value
    MontoDPTO = sheet['R11'].value
    NUnidad = sheet['S11'].value
    NTrabajadores = sheet['T11'].value
    FormaPago = sheet['V11'].value
    NCuotas = sheet['W11'].value
    FormaPago2 = sheet['X11'].value
    Banco = sheet['Y11'].value
    NroCuotas = sheet['Z11'].value
    NroTarjeta = sheet['AA11'].value
    DiaPago = sheet['AB11'].value
    Minuta = sheet['AC11'].value
    TipoTarjeta = sheet['AD11'].value
    VencimientoTarjeta = sheet['AE11'].value
    Minuta = sheet['AC11'].value
    NroCuenta = sheet['AH11'].value

    # Ingreso de rut empresa
    chrome_driver.find_element_by_xpath("//input[@id='PerAsegurado_Identificacion']").clear()
    chrome_driver.find_element_by_xpath("//input[@id='PerAsegurado_Identificacion']").send_keys(RutCondominio)
    chrome_driver.find_element_by_xpath("//input[@id='PerAsegurado_Identificacion']").send_keys(Keys.RETURN)
    sleep(1)
    chrome_driver.find_element_by_xpath("//input[@id='PerAsegurado_Identificacion']").send_keys(Keys.TAB)
    sleep(2)

    # Ingreso de razon social
    elemRazonSocial = chrome_driver.find_element_by_xpath("//input[@id='PerAsegurado_RazonSocial']")
    if(elemRazonSocial.get_attribute('value') == ''):
        elemRazonSocial.clear()
        elemRazonSocial.send_keys(RazonSocial)
        sleep(1)
    else:

        print(elemRazonSocial.get_attribute('value'))
        sleep(2)

    # Ingreso de Nro Reserva BCI
    chrome_driver.find_element_by_id("NReservaBCI_Texto").location_once_scrolled_into_view
    chrome_driver.find_element_by_id("NReservaBCI_Texto").click()
    chrome_driver.find_element_by_id("NReservaBCI_Texto").clear()
    chrome_driver.find_element_by_id("NReservaBCI_Texto").send_keys(NroReserva)
    # Ingreso de Nro Reserva Consorcio
    chrome_driver.find_element_by_id("NReservaConsorcio_Texto").click()
    chrome_driver.find_element_by_id("NReservaConsorcio_Texto").clear()
    chrome_driver.find_element_by_id("NReservaConsorcio_Texto").send_keys(CONSORCIO)
    # Ingreso de Nro Reserva Chubb
    chrome_driver.find_element_by_id("NReservaChubb_Texto").click()
    chrome_driver.find_element_by_id("NReservaChubb_Texto").clear()
    chrome_driver.find_element_by_id("NReservaChubb_Texto").send_keys(CHUBB)
    # Ingreso de Nro Reserva Renta
    chrome_driver.find_element_by_id("NReservaRenta_Texto").click()
    chrome_driver.find_element_by_id("NReservaRenta_Texto").clear()
    chrome_driver.find_element_by_id("NReservaRenta_Texto").send_keys(Renta)
    # Ingreso de Direccion
    chrome_driver.find_element_by_id("Direccion_Texto").click()
    chrome_driver.find_element_by_id("Direccion_Texto").clear()
    chrome_driver.find_element_by_id("Direccion_Texto").send_keys(Direccion)

    wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//div[@id='ui-id-2']")))
    chrome_driver.find_element_by_xpath("//div[@id='ui-id-2']").click()
    sleep(1)
    chrome_driver.save_screenshot('..\Screenshot\CP10\Parte01_FormularioCondominio.png')
    # Seleccion tipo de construccion
    chrome_driver.find_element_by_xpath("//span[@id='select2-TipoConstruccionMuro-container']").location_once_scrolled_into_view
    chrome_driver.find_element_by_xpath("//span[@id='select2-TipoConstruccionMuro-container']").click()
    sleep(1)
    chrome_driver.find_element_by_xpath("(//input[@type='search'])[2]").send_keys(ConstruccionMuro)
    sleep(1)
    chrome_driver.find_element_by_xpath("(//input[@type='search'])[2]").send_keys(Keys.ENTER)
    sleep(1)
    # Seleccion tipo de construccion techo
    chrome_driver.find_element_by_xpath("//span[@id='select2-TipoConstruccionTecho-container']").click()
    sleep(1)
    chrome_driver.find_element_by_xpath("(//input[@type='search'])[2]").send_keys(ConstTecho)
    sleep(1)
    chrome_driver.find_element_by_xpath("(//input[@type='search'])[2]").send_keys(Keys.ENTER)
    sleep(1)
    # Seleccion tipo de condominio
    chrome_driver.find_element_by_xpath("//span[@id='select2-Tipo_de_Condominio_TablaSimple_Texto-container']").click()
    sleep(1)
    chrome_driver.find_element_by_xpath("(//input[@type='search'])[2]").send_keys(TipoCondominio)
    sleep(1)
    chrome_driver.find_element_by_xpath("(//input[@type='search'])[2]").send_keys(Keys.ENTER)
    sleep(1)
    # Ingreso de Nro de pisos
    chrome_driver.find_element_by_xpath("//input[@id='NroDePisos_Text_Texto_Entero']").click()
    chrome_driver.find_element_by_xpath("//input[@id='NroDePisos_Text_Texto_Entero']").clear()
    chrome_driver.find_element_by_xpath("//input[@id='NroDePisos_Text_Texto_Entero']").send_keys(NPisos)
    # Ingreso de Nro de subterraneos
    chrome_driver.find_element_by_xpath("//input[@id='NroDeSubterraneos_Texto_Entero']").click()
    chrome_driver.find_element_by_xpath("//input[@id='NroDeSubterraneos_Texto_Entero']").clear()
    chrome_driver.find_element_by_xpath("//input[@id='NroDeSubterraneos_Texto_Entero']").send_keys(NSubterraneos)
    chrome_driver.save_screenshot('..\Screenshot\CP10\Parte02_FormularioCondominio.png')
    # Ingreso de monto de bienes
    chrome_driver.find_element_by_xpath("//input[@id='MontoBienes_Texto']").location_once_scrolled_into_view
    chrome_driver.find_element_by_xpath("//input[@id='MontoBienes_Texto']").click()
    chrome_driver.find_element_by_xpath("//input[@id='MontoBienes_Texto']").clear()
    chrome_driver.find_element_by_xpath("//input[@id='MontoBienes_Texto']").send_keys(BienesEspacio)
    # Ingreso de monto de dptos y oficinas
    #chrome_driver.find_element_by_xpath("//input[@id='MontoDeptoOficinas_Texto']").click()
    #chrome_driver.find_element_by_xpath("//input[@id='MontoDeptoOficinas_Texto']").clear()
    #chrome_driver.find_element_by_xpath("//input[@id='MontoDeptoOficinas_Texto']").send_keys(MontoDPTO)
    # Ingreso de cantidad de unidades
    #chrome_driver.find_element_by_xpath("//input[@id='CantUnidades_Texto']").click()
    #chrome_driver.find_element_by_xpath("//input[@id='CantUnidades_Texto']").clear()
    #chrome_driver.find_element_by_xpath("//input[@id='CantUnidades_Texto']").send_keys(NUnidad)
    # Ingreso de Nro de trabajadores
    chrome_driver.find_element_by_xpath("//input[@id='NumTrabajadores_Texto']").click()
    chrome_driver.find_element_by_xpath("//input[@id='NumTrabajadores_Texto']").clear()
    chrome_driver.find_element_by_xpath("//input[@id='NumTrabajadores_Texto']").send_keys(NTrabajadores)
    elementMedioPago = chrome_driver.find_element_by_xpath("//span[@id='select2-TipoMedioPago-container']")
    if (not elementMedioPago):
        chrome_driver.save_screenshot('..\Screenshot\CP10\Parte03_FormularioCondominio.png')
        chrome_driver.find_element_by_xpath("//div[@id='wizCondominio']/section[2]/div[6]/a[6]").click()
        sleep(2)
    else:
        # Seleccionar tipo de medio de pago
        chrome_driver.find_element_by_xpath("//span[@id='select2-TipoMedioPago-container']").click()
        sleep(1)
        chrome_driver.find_element_by_xpath("(//input[@type='search'])[2]").send_keys(FormaPago)
        sleep(1)
        chrome_driver.find_element_by_xpath("(//input[@type='search'])[2]").send_keys(Keys.ENTER)
        sleep(1)
        # Seleccionar el Nro de cuotas
        chrome_driver.find_element_by_xpath("//span[@id='select2-Cuotas-container']").click()
        sleep(1)
        chrome_driver.find_element_by_xpath("(//input[@type='search'])[2]").send_keys(NCuotas)
        sleep(1)
        chrome_driver.find_element_by_xpath("(//input[@type='search'])[2]").send_keys(Keys.ENTER)
        sleep(1)
        # Enviar formulario
        chrome_driver.save_screenshot('..\Screenshot\CP10\Parte03_FormularioCondominio.png')
        chrome_driver.find_element_by_xpath("//div[@id='wizCondominio']/section[2]/div[6]/a[6]").click()
        sleep(2)

    chrome_driver.save_screenshot('..\Screenshot\CP10\Envio_FormularioCondominio.png')

    #Cierra mensaje de alerta
    wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "(//a[@onclick='closeAlert(tarificarUiJs.mostrarMensajeEplacement)'])[4]")))
    chrome_driver.save_screenshot('..\Screenshot\CP10\Cierre_MensajeAlerta.png')
    sleep(2)
    chrome_driver.find_element_by_xpath("(//a[@onclick='closeAlert(tarificarUiJs.mostrarMensajeEplacement)'])[4]").click()
    sleep(2)
    #print(chrome_driver.find_element_by_xpath("//div[@id='Tarificacion.Dto']/div[2]/div/label/span").get_attribute('value'))
    #chrome_driver.find_element_by_xpath("//button[@type='button']").send_keys('CHUBB')
    chrome_driver.find_element_by_xpath("//div[@id='Plan-63']").click()
    sleep(5)
    chrome_driver.save_screenshot('..\Screenshot\CP10\Seleccion_Plan.png')
    sleep(2)
    wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//div[@id='wizCondominio']/section[2]/div[6]/a[9]")))
    chrome_driver.find_element_by_xpath("//div[@id='wizCondominio']/section[2]/div[6]/a[9]").click()
    wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//div[@id='TarificadorSeleccion']/div[2]/div/div/div/div/ul/li/label/span")))
    chrome_driver.find_element_by_xpath("//div[@id='TarificadorSeleccion']/div[2]/div/div/div/div/ul/li/label/span").click()
    chrome_driver.save_screenshot('..\Screenshot\CP10\Seleccion_Plan_Incendio.png')
    sleep(2)
    chrome_driver.find_element_by_xpath("//div[3]/div[2]/div/div/div/div/button").send_keys(Keys.ENTER)
    wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//div[@id='stpCompletarDatos']/div[3]/article")))
    sleep(1)
    chrome_driver.save_screenshot('..\Screenshot\CP10\Paso3_Parte1_CompletarDatos.png')
    sleep(1)
    #Ingreso de información en campo minuta
    chrome_driver.find_element_by_xpath("//textarea[@name='minuta']").location_once_scrolled_into_view
    chrome_driver.find_element_by_xpath("//textarea[@name='minuta']").click()
    sleep(1)
    chrome_driver.find_element_by_xpath("//textarea[@name='minuta']").send_keys(Minuta)
    sleep(1)
    chrome_driver.find_element_by_xpath("//textarea[@name='minuta']").send_keys(Keys.ENTER)
    sleep(1)
    chrome_driver.save_screenshot('..\Screenshot\CP10\Paso3_Parte2_CompletarDatos.png')
    sleep(1)

    #Seleccionar Medio de pago
    chrome_driver.find_element_by_xpath("//span[@id='select2-FormaPagoModel_formapago-container']").location_once_scrolled_into_view
    chrome_driver.find_element_by_xpath("//span[@id='select2-FormaPagoModel_formapago-container']").click()
    sleep(1)
    chrome_driver.find_element_by_xpath("//input[@type='search']").send_keys(FormaPago2)
    sleep(1)
    chrome_driver.find_element_by_xpath("//input[@type='search']").send_keys(Keys.ENTER)

    #Calcular cuotas
    chrome_driver.find_element_by_xpath("//a[@onclick='onClickRecalcularCuotas();']").click()
    sleep(1)
    chrome_driver.save_screenshot('..\Screenshot\CP10\CalculoCuota.png')
    # Scroll hacia el boton guardar
    chrome_driver.find_element_by_xpath("//div[@id='wizCondominio']/section[2]/div[6]/a[5]").location_once_scrolled_into_view
    sleep(1)
    chrome_driver.save_screenshot('..\Screenshot\CP10\Paso3_Parte3_CompletarDatos.png')
    sleep(1)
    chrome_driver.find_element_by_xpath("//div[@id='wizCondominio']/section[2]/div[6]/a[5]").click()
    #Validacion que cargue boton emitir
    wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//div[@id='wizCondominio']/section[2]/div[6]/a[7]")))
    #Scroll hacia boton emitir
    chrome_driver.find_element_by_xpath("//div[@id='wizCondominio']/section[2]/div[6]/a[7]").location_once_scrolled_into_view
    sleep(1)
    chrome_driver.save_screenshot('..\Screenshot\CP10\Paso3_Emitir poliza.png')
    sleep(1)
    # Click en boton emitir
    chrome_driver.find_element_by_xpath("//div[@id='wizCondominio']/section[2]/div[6]/a[7]").click()
    #Validación de elemento exisitente si carga interfaz final de descarga
    wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//div[@id='stpEmitir']/div[2]/div/p")))
    sleep(1)
    chrome_driver.save_screenshot('..\Screenshot\CP10\PasoFinalDescarga.png')
    sleep(1)
    #Descargar poliza
    chrome_driver.find_element_by_xpath("//a[@id='btnDownloadPoliza']/span").click()
    sleep(60)
    chrome_driver.save_screenshot('..\Screenshot\CP10\ValidacionDescarga.png')


    chrome_driver.close()


if __name__ == "__main__":
    # if __name__ == '__main__':
    unittest.main(testRunner=HtmlTestRunner.HTMLTestRunner(output= '../reports'))