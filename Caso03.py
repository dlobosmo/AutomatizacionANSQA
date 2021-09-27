# Implementation of Selenium WebDriver with Python using PyTest
import unittest
import pytest
from selenium import webdriver
import sys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.by import By
from time import sleep
import HtmlTestRunner
# importamos el submodulo "Workbook"
from openpyxl import load_workbook


def test03_condominio():
    # Especificamos el nombre y la ruta del archivo de datos a leer
    filesheet = "..\Datos\Datos_Condominio.xlsx"

    # Creamos el obejeto load_workbook
    wb = load_workbook(filesheet)

    # seleccionamos la Hoja del archivo
    sheet = wb['User']
    urlPagina = sheet['D2'].value
    rutaChromeDriver = "..\drivers\chromedriver.exe"
    chrome_driver = webdriver.Chrome(rutaChromeDriver)

    chrome_driver.get(urlPagina)
    chrome_driver.maximize_window()


    # Obtenemos el valor de la celda para leer el usuario y contraseña de quien iniciará sesion en el sistema
    email = sheet['A2'].value
    passw = sheet['B2'].value

    wait = WebDriverWait(chrome_driver, 70)
    wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//input[@id='login-password']")))

    # Se ingresa el usuario del sistema
    elem = chrome_driver.find_element_by_xpath("//input[@id='login-user-name']")
    elem.clear()
    elem.send_keys(email)

    # Se ingresa la password del usuario del sistema
    elem = chrome_driver.find_element_by_xpath("//input[@id='login-password']")
    elem.clear()
    elem.send_keys(passw)
    chrome_driver.save_screenshot('..\Screenshot\CP03\Inicio_Sesion.png')
    elem.send_keys(Keys.RETURN)

    # Voy a pinchar opción Mantenedores en el menú
    wait = WebDriverWait(chrome_driver, 120)
    wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//span[text()='Mantenedores']")))
    elem = chrome_driver.find_element_by_xpath("//span[text()='Mantenedores']").click()
    # Acá llama a la opción Ramo Wizard pero con menú extendido
    elem = chrome_driver.find_element_by_xpath("//a[text()='Condominio']").click()
    chrome_driver.save_screenshot('..\Screenshot\CP03\Home_Sistema.png')

    # sleep(10)
    timeout = 30
    try:
        element_present = expected_conditions.presence_of_element_located((By.ID, 'iframe-render'))
        WebDriverWait(chrome_driver, timeout).until(element_present)
    except TimeoutException:
        print('Timed out waiting for page to load')

    # Me cambio al iframe de ingreso de datos
    iframe = chrome_driver.find_element_by_id("iframe-render")
    chrome_driver.switch_to.frame(iframe)

    # seleccionamos la Hoja del archivo con los datos de prueba del caso
    sheet = wb['Datos CP 1']

    RutCondominio = sheet['B4'].value
    RazonSocial = sheet['C4'].value
    NroReserva = sheet['D4'].value
    CONSORCIO = sheet['E4'].value
    CHUBB = sheet['F4'].value
    Renta = sheet['G4'].value
    Comuna = sheet['H4'].value
    Direccion = sheet['I4'].value
    ConstruccionMuro = sheet['J4'].value
    ConstTecho = sheet['K4'].value
    TipoCondominio = sheet['L4'].value
    NPisos = sheet['M4'].value
    NSubterraneos = sheet['N4'].value
    Antiguedad = sheet['O4'].value
    UbicacionRiesgo = sheet['P4'].value
    BienesEspacio = sheet['Q4'].value
    MontoDPTO = sheet['R4'].value
    NUnidad = sheet['S4'].value
    NTrabajadores = sheet['T4'].value
    FormaPago = sheet['V4'].value
    NCuotas = sheet['W4'].value

    # Ingreso de rut empresa
    chrome_driver.find_element_by_xpath("//input[@id='PerAsegurado_Identificacion']").clear()
    chrome_driver.find_element_by_xpath("//input[@id='PerAsegurado_Identificacion']").send_keys(RutCondominio)
    chrome_driver.find_element_by_xpath("//input[@id='PerAsegurado_Identificacion']").send_keys(Keys.RETURN)
    sleep(1)
    chrome_driver.find_element_by_xpath("//input[@id='PerAsegurado_Identificacion']").send_keys(Keys.TAB)
    # Ingreso de razon social
    elemRazonSocial = chrome_driver.find_element_by_xpath("//input[@id='PerAsegurado_RazonSocial']")
    if (elemRazonSocial.get_attribute('value') == ''):
        elemRazonSocial.clear()
        elemRazonSocial.send_keys(RazonSocial)
        sleep(1)
    else:
        #assert elemRazonSocial.get_attribute('value') == RazonSocial
        print(elemRazonSocial.get_attribute('value'))
        sleep(2)
    # Ingreso de Nro Reserva BCI
    chrome_driver.find_element_by_id("NReservaBCI_Texto").location_once_scrolled_into_view
    chrome_driver.find_element_by_id("NReservaBCI_Texto").click()
    chrome_driver.find_element_by_id("NReservaBCI_Texto").clear()
    chrome_driver.find_element_by_id("NReservaBCI_Texto").send_keys(NroReserva)
    # Ingreso de Nro Reserva Consorcio
    chrome_driver.find_element_by_id("NReservaConsorcio_Texto").click()
    chrome_driver.find_element_by_id("NReservaConsorcio_Texto").clear()
    chrome_driver.find_element_by_id("NReservaConsorcio_Texto").send_keys(CONSORCIO)
    # Ingreso de Nro Reserva Chubb
    chrome_driver.find_element_by_id("NReservaChubb_Texto").click()
    chrome_driver.find_element_by_id("NReservaChubb_Texto").clear()
    chrome_driver.find_element_by_id("NReservaChubb_Texto").send_keys(CHUBB)
    # Ingreso de Nro Reserva Renta
    chrome_driver.find_element_by_id("NReservaRenta_Texto").click()
    chrome_driver.find_element_by_id("NReservaRenta_Texto").clear()
    chrome_driver.find_element_by_id("NReservaRenta_Texto").send_keys(Renta)
    # Ingreso de Direccion
    chrome_driver.find_element_by_id("Direccion_Texto").click()
    chrome_driver.find_element_by_id("Direccion_Texto").clear()
    chrome_driver.find_element_by_id("Direccion_Texto").send_keys(Direccion)
    wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//div[@id='ui-id-2']")))
    chrome_driver.find_element_by_xpath("//div[@id='ui-id-2']").click()
    sleep(1)
    chrome_driver.save_screenshot('..\Screenshot\CP03\Parte01_FormularioCondominio.png')
    # Seleccion tipo de construccion
    chrome_driver.find_element_by_xpath("//span[@id='select2-TipoConstruccionMuro-container']").location_once_scrolled_into_view
    chrome_driver.find_element_by_xpath("//span[@id='select2-TipoConstruccionMuro-container']").click()
    sleep(1)
    chrome_driver.find_element_by_xpath("(//input[@type='search'])[2]").send_keys(ConstruccionMuro)
    sleep(1)
    chrome_driver.find_element_by_xpath("(//input[@type='search'])[2]").send_keys(Keys.ENTER)
    sleep(1)
    # Seleccion tipo de construccion techo
    chrome_driver.find_element_by_xpath("//span[@id='select2-TipoConstruccionTecho-container']").click()
    sleep(1)
    chrome_driver.find_element_by_xpath("(//input[@type='search'])[2]").send_keys(ConstTecho)
    sleep(1)
    chrome_driver.find_element_by_xpath("(//input[@type='search'])[2]").send_keys(Keys.ENTER)
    sleep(1)
    # Seleccion tipo de condominio
    chrome_driver.find_element_by_xpath("//span[@id='select2-Tipo_de_Condominio_TablaSimple_Texto-container']").click()
    sleep(1)
    chrome_driver.find_element_by_xpath("(//input[@type='search'])[2]").send_keys(TipoCondominio)
    sleep(1)
    chrome_driver.find_element_by_xpath("(//input[@type='search'])[2]").send_keys(Keys.ENTER)
    sleep(1)
    # Ingreso de Nro de pisos
    chrome_driver.find_element_by_xpath("//input[@id='NroDePisos_Text_Texto_Entero']").click()
    chrome_driver.find_element_by_xpath("//input[@id='NroDePisos_Text_Texto_Entero']").clear()
    chrome_driver.find_element_by_xpath("//input[@id='NroDePisos_Text_Texto_Entero']").send_keys(NPisos)
    # Ingreso de Nro de subterraneos
    chrome_driver.find_element_by_xpath("//input[@id='NroDeSubterraneos_Texto_Entero']").click()
    chrome_driver.find_element_by_xpath("//input[@id='NroDeSubterraneos_Texto_Entero']").clear()
    chrome_driver.find_element_by_xpath("//input[@id='NroDeSubterraneos_Texto_Entero']").send_keys(NSubterraneos)
    chrome_driver.save_screenshot('..\Screenshot\CP03\Parte02_FormularioCondominio.png')
    # Ingreso de monto de bienes
    chrome_driver.find_element_by_xpath("//input[@id='MontoBienes_Texto']").location_once_scrolled_into_view
    chrome_driver.find_element_by_xpath("//input[@id='MontoBienes_Texto']").click()
    chrome_driver.find_element_by_xpath("//input[@id='MontoBienes_Texto']").clear()
    chrome_driver.find_element_by_xpath("//input[@id='MontoBienes_Texto']").send_keys(BienesEspacio)

    # Ingreso de monto de dptos y oficinas
    #chrome_driver.find_element_by_xpath("//input[@id='MontoDeptoOficinas_Texto']").click()
    #chrome_driver.find_element_by_xpath("//input[@id='MontoDeptoOficinas_Texto']").clear()
    #chrome_driver.find_element_by_xpath("//input[@id='MontoDeptoOficinas_Texto']").send_keys(MontoDPTO)
    # Ingreso de cantidad de unidades
    #chrome_driver.find_element_by_xpath("//input[@id='CantUnidades_Texto']").click()
    #chrome_driver.find_element_by_xpath("//input[@id='CantUnidades_Texto']").clear()
    #chrome_driver.find_element_by_xpath("//input[@id='CantUnidades_Texto']").send_keys(NUnidad)

    # Ingreso de Nro de trabajadores
    chrome_driver.find_element_by_xpath("//input[@id='NumTrabajadores_Texto']").click()
    chrome_driver.find_element_by_xpath("//input[@id='NumTrabajadores_Texto']").clear()
    chrome_driver.find_element_by_xpath("//input[@id='NumTrabajadores_Texto']").send_keys(NTrabajadores)

    elementMedioPago = chrome_driver.find_element_by_xpath("//span[@id='select2-TipoMedioPago-container']")
    if (not elementMedioPago):
        chrome_driver.save_screenshot('..\Screenshot\CP03\Parte03_FormularioCondominio.png')
        chrome_driver.find_element_by_xpath("//div[@id='wizCondominio']/section[2]/div[6]/a[6]").click()
        sleep(2)
    else:
        # Seleccionar tipo de medio de pago
        chrome_driver.find_element_by_xpath("//span[@id='select2-TipoMedioPago-container']").click()
        sleep(1)
        chrome_driver.find_element_by_xpath("(//input[@type='search'])[2]").send_keys(FormaPago)
        sleep(1)
        chrome_driver.find_element_by_xpath("(//input[@type='search'])[2]").send_keys(Keys.ENTER)
        sleep(1)
        # Seleccionar el Nro de cuotas
        chrome_driver.find_element_by_xpath("//span[@id='select2-Cuotas-container']").click()
        sleep(1)
        chrome_driver.find_element_by_xpath("(//input[@type='search'])[2]").send_keys(NCuotas)
        sleep(1)
        chrome_driver.find_element_by_xpath("(//input[@type='search'])[2]").send_keys(Keys.ENTER)
        sleep(1)
        # Enviar formulario
        chrome_driver.save_screenshot('..\Screenshot\CP03\Parte03_FormularioCondominio.png')
        chrome_driver.find_element_by_xpath("//div[@id='wizCondominio']/section[2]/div[6]/a[6]").click()
        sleep(2)

    chrome_driver.save_screenshot('..\Screenshot\CP03\Envio_FormularioCondominio.png')

    #Cierra mensaje de alerta
    wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "(//a[@onclick='closeAlert(tarificarUiJs.mostrarMensajeEplacement)'])[4]")))
    chrome_driver.save_screenshot('..\Screenshot\CP03\Cierre_MensajeAlerta.png')
    sleep(2)
    chrome_driver.find_element_by_xpath("(//a[@onclick='closeAlert(tarificarUiJs.mostrarMensajeEplacement)'])[4]").click()
    sleep(2)
    print(chrome_driver.find_element_by_xpath("//div[@id='Tarificacion.Dto']/div[2]/div/label/span").get_attribute('value'))
    chrome_driver.find_element_by_xpath("//div[@id='toolsActionsOferta']/div[2]/span").click()
    sleep(20)
    chrome_driver.save_screenshot('..\Screenshot\CP03\Documento_PDF.png')
    chrome_driver.close()


if __name__ == "__main__":
    # if __name__ == '__main__':
    unittest.main(testRunner=HtmlTestRunner.HTMLTestRunner(output= '../reports'))